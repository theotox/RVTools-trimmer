#!/usr/bin/env python3
"""
clean_rvtools.py - Strip RVtools xlsx reports down to sizing-relevant columns.

Usage:
    python clean_rvtools.py report1.xlsx [report2.xlsx ...]

Output:
    <name>_cleaned.xlsx next to each input file.
"""

import argparse
import os
import sys

from openpyxl import Workbook, load_workbook

# Per-sheet column keep lists.
# If a listed column header isn't found, it's silently skipped.
# Any sheet not listed here is omitted entirely from the output.
KEEP_COLUMNS: dict[str, list[str]] = {
    # --- vInfo: consolidated VM overview (95 cols -> 14) ---
    "vInfo": [
        "VM",
        "Powerstate",
        "CPUs",
        "Memory",
        "Active Memory",
        "NICs",
        "Disks",
        "Total disk capacity MiB",
        "Provisioned MiB",
        "In Use MiB",
        "Resource pool",
        "Folder",
        "Primary IP Address",
    ],
    # --- vCPU: CPU sizing (34 cols -> 11) ---
    "vCPU": [
        "VM",
        "Powerstate",
        "CPUs",
        "Sockets",
        "Cores p/s",
        "Reservation",
        "Entitlement",
        "DRS Entitlement",
        "Cluster",
        "Host",
        "OS according to the VMware Tools",
    ],
    # --- vMemory: Memory sizing (38 cols -> 5) ---
    "vMemory": [
        "VM",
        "Size MiB",
        "Active",
        "Cluster",
        "Host",
    ],
    # --- vDisk: Disk sizing (45 cols -> 9) ---
    "vDisk": [
        "VM",
        "Disk",
        "Capacity MiB",
        "Thin",
        "Disk Mode",
        "Sharing mode",
        "Controller",
        "Cluster",
        "Host",
    ],
    # --- vNetwork: Network (32 cols -> 8) ---
    "vNetwork": [
        "VM",
        "NIC label",
        "Adapter",
        "Network",
        "Connected",
        "Mac Address",
        "Type",
        "IPv4 Address",
    ],
    # --- vCluster: Clusters (35 cols -> 11) ---
    "vCluster": [
        "Name",
        "NumHosts",
        "TotalCpu",
        "NumCpuCores",
        "NumCpuThreads",
        "TotalMemory",
        "HA enabled",
        "AdmissionControlEnabled",
        "DRS enabled",
        "VI SDK Server",
    ],
    # --- vHost: Hosts (74 cols -> 16) ---
    "vHost": [
        "Host",
        "Datacenter",
        "Cluster",
        "CPU Model",
        "Speed",
        "# CPU",
        "Cores per CPU",
        "# Cores",
        "# Memory",
        "# VMs total",
        "# VMs",
        "# vCPUs",
        "vCPUs per Core",
        "vRAM",
        "ESX Version",
        "VMotion support",
    ],
    # --- vDatastore: Datastores (32 cols -> 10) ---
    "vDatastore": [
        "Name",
        "Type",
        "Capacity MiB",
        "Provisioned MiB",
        "In Use MiB",
        "Free MiB",
        "Free %",
        "# VMs",
        "# Hosts",
        "Cluster name",
    ],
    # --- vSnapshot: Snapshots (26 cols -> 6) ---
    "vSnapshot": [
        "VM",
        "Name",
        "Date / time",
        "Size MiB (total)",
        "State",
        "Cluster",
    ],
    # --- vTools: VM Tools (34 cols -> 6) ---
    "vTools": [
        "VM",
        "VM Version",
        "Tools",
        "Tools Version",
        "Upgradeable",
        "Cluster",
    ],
    # === Legacy RVtools naming (older versions / different export options) ===
    "VMs": [
        "VM",
        "Powerstate",
        "CPUs",
        "Sockets",
        "Cores p/s",
        "Reservation",
        "Entitlement",
        "DRS Entitlement",
        "Cluster",
        "Host",
        "OS according to the VMware Tools",
    ],
    "VMDisks": [
        "VM",
        "Disk",
        "Provisioned Size",
        "Uncommitted",
        "Shared",
        "Disk Type",
        "Datastore",
        "Thin Provisioned",
    ],
    "VMNetwork": [
        "VM",
        "Network Adapter",
        "Type",
        "Network Label",
        "Connected",
        "Connection Type",
    ],
    "VMemory": [
        "VM",
        "Size MiB",
        "Active",
    ],
    "Hosts": [
        "DNS Name",
        "Version",
        "CPUs",
        "CPU Model",
        "CPU Threads",
        "CPU Speed",
        "Memory Total",
        "Memory Used",
        "Cluster",
        "Datacenter",
        "vCPUs Total",
        "vCPUs Used",
        "VMs",
    ],
    "Clusters": [
        "Name",
        "Datacenter",
        "Hosts",
        "VMs",
        "vCPUs",
        "RAM",
        "DRS Enabled",
        "HA Enabled",
    ],
    "Datastores": [
        "Name",
        "Type",
        "Capacity",
        "Free Space",
        "VMs",
        "Hosts",
        "Datacenter",
    ],
    "VMHardware": [
        "VM",
        "Hardware Version",
    ],
    "VMTools": [
        "VM",
        "VMware Tools version",
        "VMware Tools version number",
    ],
    "Snapshots": [
        "VM",
        "Created",
        "Days",
    ],
    "HostHardware": [
        "DNS Name",
        "CPUs",
        "CPU Model",
        "CPU Threads",
        "CPU Speed",
        "Memory Total",
    ],
}


def clean_sheet(ws, wb_out: Workbook, sheet_name: str, keep: list[str]) -> None:
    """Copy only the columns in *keep* from *ws* into a new sheet in *wb_out*."""
    # Read header row (first row of the read-only worksheet)
    header_cells = []
    for cell in ws[1]:
        header_cells.append(cell.value)

    if not header_cells:
        print(f"  [skip] '{sheet_name}' is empty")
        return

    new_ws = wb_out.create_sheet(title=sheet_name)

    # Locate column indices for the keep list
    col_map: list[tuple[int, str]] = []
    for wanted in keep:
        try:
            idx = header_cells.index(wanted) + 1
            col_map.append((idx, wanted))
        except ValueError:
            print(f"  [warn] column '{wanted}' not found in '{sheet_name}', skipping")

    if not col_map:
        print(f"  [warn] no matching columns in '{sheet_name}', skipping")
        wb_out.remove(new_ws)
        return

    # Copy header
    for new_col, (_, wanted) in enumerate(col_map, start=1):
        new_ws.cell(row=1, column=new_col, value=wanted)

    # Copy data rows
    row_count = 0
    for row in ws.iter_rows(min_row=2):
        row_count += 1
        for new_col, (src_col, _) in enumerate(col_map, start=1):
            out = new_ws.cell(row=row_count + 1, column=new_col)
            out.value = row[src_col - 1].value

    print(f"  [ok] '{sheet_name}': {len(col_map)} cols, {row_count} rows")


def clean_file(in_path: str) -> None:
    base = os.path.splitext(in_path)[0]
    out_path = f"{base}_cleaned.xlsx"

    print(f"\nProcessing: {in_path}")
    wb_in = load_workbook(in_path, read_only=True, data_only=True)

    wb_out = Workbook()
    del wb_out["Sheet"]

    sheets_kept = 0
    for sheet_name in KEEP_COLUMNS:
        if sheet_name in wb_in.sheetnames:
            clean_sheet(wb_in[sheet_name], wb_out, sheet_name, KEEP_COLUMNS[sheet_name])
            sheets_kept += 1
        else:
            print(f"  [skip] sheet '{sheet_name}' not in workbook")

    wb_in.close()

    if sheets_kept == 0:
        print("[error] no sheets matched — aborting. Check KEEP_COLUMNS in script.")
        return

    wb_out.save(out_path)
    print(f"\nSaved: {out_path}  ({sheets_kept} sheets)")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Clean RVtools xlsx — keep only sizing-relevant columns."
    )
    parser.add_argument("files", nargs="+", help="RVtools .xlsx file(s) to clean")
    args = parser.parse_args()

    for fpath in args.files:
        if not os.path.isfile(fpath):
            print(f"Error: {fpath} not found", file=sys.stderr)
            sys.exit(1)
        clean_file(fpath)


if __name__ == "__main__":
    main()
